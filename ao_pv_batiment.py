"""
Module AO PV Bâtiment - CRE PPE2
Réponse automatique à l'appel d'offres CRE pour centrales PV sur bâtiments > 500 kWc
Version : février 2026 (Période 12 - date limite : vendredi 24 avril 2026)

Fonctionnalités :
- Saisie guidée wizard multi-étapes à partir des projets/prospects CRM
- Simulation des notes NP (prix), NC (carbone), NFC (gouvernance/financement)
- Calcul du bilan carbone simplifié (Annexe 2 CDC)
- Génération PDF du formulaire pré-rempli
- Export Excel du tableau de bilan carbone
"""

from flask import Blueprint, render_template, request, jsonify, send_file, session
from datetime import datetime
import json
import math
import io

ao_pv_bp = Blueprint('ao_pv', __name__, url_prefix='/ao-pv-batiment')

# ─── DONNÉES DE RÉFÉRENCE AO PPE2 PV Bâtiment ─────────────────────────────────

# Prix plafonds par période (€/MWh) - périodes 1-3 publiées, reste confidentiel
PRIX_PLAFOND = {
    1: 96, 2: 96, 3: 96,
    # 4-14: confidentiel, on utilise l'estimation de marché
    **{i: None for i in range(4, 15)}
}
PERIODE_ACTUELLE = 12
DATE_LIMITE_P12 = "24 avril 2026"

# Bilans carbone plafond/plancher par période (kg eq CO2/kWc)
CARBONE_REF = {
    **{i: {"sup": 550, "inf": 200} for i in range(1, 8)},
    **{i: {"sup": 740, "inf": 420} for i in range(8, 15)},
}

# Pondérations de notation (§4.1)
POIDS = {"NP": 70, "NC": 25, "NFC": 5}  # GP max 5, FC max 2

# ─── TABLEAU 3 : GWP PAR PAYS (extrait CDC Annexe 2) ──────────────────────────
# Format : {pays: {etape: valeur_gwp}}
# Unités : kg CO2-eq/kg ou kg CO2-eq/m² selon l'étape

GWP_PAR_PAYS = {
    "France":     {"mg_si": 7.34, "poly_si": 27.23, "lingot_mono": 15.39, "lingot_multi": 2.20, "lingot_monolike": 4.82, "brique": 0.82, "wafer_mono": 3.54, "wafer_multi": 3.95, "cellule": 19.63, "verre": 1.00, "verre_trempe": 0.07, "encapsulant": 2.67, "face_arriere": 3.69, "pvf": 20.43, "module_cristallin": 4.82, "module_asi": 23.47, "module_cdte": 23.48, "module_cigs": 35.38},
    "Allemagne":  {"mg_si": 12.25, "poly_si": 58.51, "lingot_mono": 29.68, "lingot_multi": 5.32, "lingot_monolike": 7.95, "brique": 1.36, "wafer_mono": 5.76, "wafer_multi": 6.15, "cellule": 30.82, "verre": 1.05, "verre_trempe": 0.07, "encapsulant": 2.97, "face_arriere": 3.99, "pvf": 21.68, "module_cristallin": 6.51, "module_asi": 45.01, "module_cdte": 49.45, "module_cigs": 90.06},
    "Chine":      {"mg_si": 15.37, "poly_si": 75.21, "lingot_mono": 38.77, "lingot_multi": 7.64, "lingot_monolike": 9.93, "brique": 1.71, "wafer_mono": 7.37, "wafer_multi": 7.09, "cellule": 37.91, "verre": 1.08, "verre_trempe": 0.059, "encapsulant": 3.45, "face_arriere": 4.14, "pvf": 21.97, "module_cristallin": 7.57, "module_asi": 58.68, "module_cdte": 65.92, "module_cigs": 124.75},
    "Vietnam":    {"mg_si": 12.00, "poly_si": 53.78, "lingot_mono": 28.98, "lingot_multi": 5.49, "lingot_monolike": 7.79, "brique": 1.34, "wafer_mono": 5.85, "wafer_multi": 5.59, "cellule": 30.25, "verre": 1.04, "verre_trempe": 0.055, "encapsulant": 3.24, "face_arriere": 3.93, "pvf": 21.11, "module_cristallin": 6.42, "module_asi": 43.92, "module_cdte": 48.13, "module_cigs": 87.28},
    "Malaisie":   {"mg_si": 12.56, "poly_si": 57.34, "lingot_mono": 30.60, "lingot_multi": 5.85, "lingot_monolike": 8.15, "brique": 1.40, "wafer_mono": 6.10, "wafer_multi": 5.84, "cellule": 31.52, "verre": 1.05, "verre_trempe": 0.055, "encapsulant": 3.28, "face_arriere": 3.97, "pvf": 21.26, "module_cristallin": 6.61, "module_asi": 46.37, "module_cdte": 51.08, "module_cigs": 93.50},
    "Corée du Sud": {"mg_si": 14.98, "poly_si": 72.72, "lingot_mono": 37.64, "lingot_multi": 7.39, "lingot_monolike": 9.69, "brique": 1.66, "wafer_mono": 7.19, "wafer_multi": 6.92, "cellule": 37.02, "verre": 1.07, "verre_trempe": 0.058, "encapsulant": 3.42, "face_arriere": 4.11, "pvf": 21.87, "module_cristallin": 7.44, "module_asi": 56.96, "module_cdte": 63.85, "module_cigs": 120.39},
    "Japon":      {"mg_si": 12.24, "poly_si": 55.27, "lingot_mono": 29.66, "lingot_multi": 5.64, "lingot_monolike": 7.94, "brique": 1.36, "wafer_mono": 5.95, "wafer_multi": 5.69, "cellule": 30.78, "verre": 1.04, "verre_trempe": 0.055, "encapsulant": 3.26, "face_arriere": 3.95, "pvf": 21.17, "module_cristallin": 6.50, "module_asi": 44.94, "module_cdte": 49.36, "module_cigs": 89.88},
    "Taiwan":     {"mg_si": 13.32, "poly_si": 62.18, "lingot_mono": 32.82, "lingot_multi": 6.33, "lingot_monolike": 8.63, "brique": 1.48, "wafer_mono": 6.44, "wafer_multi": 6.18, "cellule": 33.25, "verre": 1.06, "verre_trempe": 0.056, "encapsulant": 3.32, "face_arriere": 4.01, "pvf": 21.45, "module_cristallin": 6.87, "module_asi": 49.71, "module_cdte": 55.10, "module_cigs": 101.97},
    "Inde":       {"mg_si": 11.23, "poly_si": 48.87, "lingot_mono": 26.73, "lingot_multi": 5.00, "lingot_monolike": 7.30, "brique": 1.25, "wafer_mono": 5.50, "wafer_multi": 5.24, "cellule": 28.49, "verre": 1.03, "verre_trempe": 0.054, "encapsulant": 3.20, "face_arriere": 3.89, "pvf": 20.92, "module_cristallin": 6.16, "module_asi": 40.53, "module_cdte": 44.05, "module_cigs": 78.69},
    "États-Unis": {"mg_si": 12.72, "poly_si": 58.35, "lingot_mono": 31.07, "lingot_multi": 5.95, "lingot_monolike": 8.25, "brique": 1.42, "wafer_mono": 6.17, "wafer_multi": 5.91, "cellule": 31.88, "verre": 1.05, "verre_trempe": 0.056, "encapsulant": 3.29, "face_arriere": 3.98, "pvf": 21.30, "module_cristallin": 6.67, "module_asi": 47.07, "module_cdte": 51.92, "module_cigs": 95.28},
    "Suisse":     {"mg_si": 5.22, "poly_si": 13.78, "lingot_mono": 9.24, "lingot_multi": 0.85, "lingot_monolike": 3.48, "brique": 0.59, "wafer_mono": 2.59, "wafer_multi": 3.00, "cellule": 14.82, "verre": 0.98, "verre_trempe": 0.07, "encapsulant": 2.54, "face_arriere": 3.56, "pvf": 19.89, "module_cristallin": 4.10, "module_asi": 14.20, "module_cdte": 12.31, "module_cigs": 11.87},
    "Norvège":    {"mg_si": 5.18, "poly_si": 13.54, "lingot_mono": 9.12, "lingot_multi": 0.83, "lingot_monolike": 3.45, "brique": 0.59, "wafer_mono": 2.57, "wafer_multi": 2.99, "cellule": 14.73, "verre": 0.98, "verre_trempe": 0.066, "encapsulant": 2.54, "face_arriere": 3.56, "pvf": 19.88, "module_cristallin": 4.08, "module_asi": 14.03, "module_cdte": 12.11, "module_cigs": 11.43},
    "Autre pays d'Europe": {"mg_si": 8.60, "poly_si": 35.29, "lingot_mono": 19.07, "lingot_multi": 3.00, "lingot_monolike": 5.63, "brique": 0.96, "wafer_mono": 4.11, "wafer_multi": 4.51, "cellule": 22.51, "verre": 1.01, "verre_trempe": 0.07, "encapsulant": 2.75, "face_arriere": 3.77, "pvf": 20.75, "module_cristallin": 5.26, "module_asi": 29.02, "module_cdte": 30.17, "module_cigs": 49.46},
    "Autre pays du Monde": {"mg_si": 12.67, "poly_si": 58.03, "lingot_mono": 30.92, "lingot_multi": 5.92, "lingot_monolike": 8.22, "brique": 1.41, "wafer_mono": 6.15, "wafer_multi": 5.88, "cellule": 31.77, "verre": 1.05, "verre_trempe": 0.06, "encapsulant": 3.28, "face_arriere": 3.97, "pvf": 21.28, "module_cristallin": 6.65, "module_asi": 46.85, "module_cdte": 51.66, "module_cigs": 94.72},
}

# Coefficients de pertes et casses (Tableau 2 CDC)
COEFF_PERTES = {
    "poly_mg_si": 1.13,       # kg MG-Si / kg polysilicium
    "lingot_mono_poly": 1.04, # kg poly / kg lingot mono
    "lingot_multi_poly": 1.01,
    "brique_mono": 1.79,      # kg lingot / kg brique mono
    "brique_multi": 1.56,
    "encapsulant": 1.01,
    "face_arriere": 1.02,
    "verre": 1.00,
    "verre_trempe": 1.00,
    "cellule_wafer": 1.01,    # m² plaquette / m² cellule
    "module_cellule": 1.02,   # m² cellule / m² module
}


# ─── CALCUL BILAN CARBONE SIMPLIFIÉ ────────────────────────────────────────────

def calculer_bilan_carbone(config_module):
    """
    Calcule l'évaluation carbone simplifiée G (kg CO2-eq/kWc) selon Annexe 2 CDC.

    config_module : dict avec :
        - technologie : 'mono' | 'multi' | 'monolike' | 'a-si' | 'cdte' | 'cigs'
        - puissance_module_wc : puissance en Wc d'un module
        - surface_module_m2 : surface d'un module en m²
        - composants : liste de dicts :
            {nom, quantite_par_kwc, unite, sites: [{pays, fraction}]}
    """
    tech = config_module.get("technologie", "mono")
    composants = config_module.get("composants", [])
    G_total = 0.0
    details = []

    for comp in composants:
        nom = comp["nom"]
        Qi = comp["quantite_par_kwc"]
        sites = comp.get("sites", [{"pays": "Autre pays du Monde", "fraction": 1.0}])
        etape_gwp = comp.get("etape_gwp", nom)

        Gi = 0.0
        for site in sites:
            pays = site["pays"]
            xij = site["fraction"]
            gwp_data = GWP_PAR_PAYS.get(pays, GWP_PAR_PAYS["Autre pays du Monde"])
            GWPij = gwp_data.get(etape_gwp, 0)
            Gi += Qi * xij * GWPij

        G_total += Gi
        details.append({
            "composant": nom,
            "Qi": Qi,
            "Gi": round(Gi, 3),
            "sites": sites
        })

    return round(G_total, 1), details


def estimation_carbone_rapide(pays_module, pays_cellule, pays_wafer, tech="mono"):
    """
    Estimation simplifiée du bilan carbone G à partir des 3 pays principaux.
    Hypothèses standards pour un module mono 2kWc/m² (1m² = 500Wc).
    Retourne G en kg CO2-eq/kWc.
    """
    if not pays_module:
        pays_module = "Autre pays du Monde"
    if not pays_cellule:
        pays_cellule = pays_module
    if not pays_wafer:
        pays_wafer = pays_cellule

    gwp_module = GWP_PAR_PAYS.get(pays_module, GWP_PAR_PAYS["Autre pays du Monde"])
    gwp_cellule = GWP_PAR_PAYS.get(pays_cellule, GWP_PAR_PAYS["Autre pays du Monde"])
    gwp_wafer = GWP_PAR_PAYS.get(pays_wafer, GWP_PAR_PAYS["Autre pays du Monde"])
    gwp_poly = GWP_PAR_PAYS.get(pays_wafer, GWP_PAR_PAYS["Autre pays du Monde"])

    # Surface module = 500 kWc / (1000 Wc/kWc * ~0.2 rendement) ≈ 2.5 m²/kWc (module 400Wc std)
    # Quantités standards pour module 400Wc monofacial
    surface_module_kwc = 2.5   # m² par kWc (≈ 400W / m² 0.20 rendement)

    if tech in ("a-si",):
        cle_module = "module_asi"
    elif tech == "cdte":
        cle_module = "module_cdte"
    elif tech == "cigs":
        cle_module = "module_cigs"
    else:
        cle_module = "module_cristallin"

    G_module = surface_module_kwc * gwp_module.get(cle_module, 6.5)

    # Cellule ≈ 2.45 m² / kWc (avec pertes)
    surface_cellule_kwc = surface_module_kwc * COEFF_PERTES["module_cellule"]
    G_cellule = surface_cellule_kwc * gwp_cellule.get("cellule", 25.0)

    # Wafer ≈ 2.47 m² / kWc
    surface_wafer_kwc = surface_cellule_kwc * COEFF_PERTES["cellule_wafer"]
    G_wafer = surface_wafer_kwc * gwp_wafer.get(f"wafer_{tech}" if tech in ("mono","multi","monolike") else "wafer_mono", 5.0)

    # Polysilicium ≈ 2.5 kg/kWc  (wafer 160µm, ~2330 kg/m³, surface_wafer)
    masse_poly_kwc = surface_wafer_kwc * (160 + 70) * 2330 * 1e-6 / (1 - 0.0) * COEFF_PERTES["brique_mono"] * COEFF_PERTES["lingot_mono_poly"]
    G_poly = masse_poly_kwc * gwp_poly.get("poly_si", 50.0)

    # Verre : ~0.8 kg/kWc (~2kg par m², surface_module)
    masse_verre_kwc = surface_module_kwc * 0.8  # simplifié
    G_verre = masse_verre_kwc * gwp_module.get("verre", 1.0)

    # Encapsulant : ~0.5 kg/kWc
    G_encap = 0.5 * COEFF_PERTES["encapsulant"] * gwp_module.get("encapsulant", 2.7)

    G_total = G_module + G_cellule + G_wafer + G_poly + G_verre + G_encap

    # Arrondi au multiple de 10 le plus proche (règle CDC §4.3)
    G_arrondi = round(G_total / 10) * 10

    return G_arrondi, round(G_total, 1)


# ─── CALCUL DES NOTES ──────────────────────────────────────────────────────────

def calculer_note_np(prix_propose, prix_plafond=None, np0=70):
    """
    Note NP sur 70. Formule : NP = NP0 × (Psup - P) / (Psup - Pinf)
    Psup = prix_plafond, Pinf calculé dynamiquement.
    Pour simulation, on estime Pinf = Psup - 20 €/MWh.
    """
    if prix_plafond is None:
        return None, "Prix plafond confidentiel pour cette période"

    if prix_propose > prix_plafond:
        return 0, f"Prix ({prix_propose} €/MWh) > plafond ({prix_plafond} €/MWh) → offre éliminée"

    # Estimation Pinf pour simulation (en réalité calculé après dépôt de toutes les offres)
    p_inf_estim = prix_plafond - 20
    if prix_propose <= p_inf_estim:
        np = np0
        note = f"Note maximale (prix très compétitif)"
    else:
        np = np0 * (prix_plafond - prix_propose) / (prix_plafond - p_inf_estim)
        note = f"NP = {np0} × ({prix_plafond} - {prix_propose}) / ({prix_plafond} - {p_inf_estim})"

    return round(np, 2), note


def calculer_note_nc(ecs_propose, periode=12, nc0=25):
    """
    Note NC sur 25. Formule : NC = NC0 × (ECSsup - ECS) / (ECSsup - ECSinf)
    ECS en kg CO2-eq/kWc, arrondi au multiple de 10.
    """
    ref = CARBONE_REF.get(periode, {"sup": 740, "inf": 420})
    ecs_sup = ref["sup"]
    ecs_inf = ref["inf"]

    # Arrondir ECS au multiple de 10 le plus proche (§4.3 CDC)
    ecs_arrondi = round(ecs_propose / 10) * 10

    if ecs_arrondi > ecs_sup:
        return 0, f"ECS ({ecs_arrondi}) > plafond ({ecs_sup}) → offre éliminée"

    if ecs_arrondi <= ecs_inf:
        return nc0, f"Note maximale (ECS ≤ plancher {ecs_inf})"

    nc = nc0 * (ecs_sup - ecs_arrondi) / (ecs_sup - ecs_inf)
    detail = f"NC = {nc0} × ({ecs_sup} - {ecs_arrondi}) / ({ecs_sup} - {ecs_inf})"

    return round(nc, 2), detail


def calculer_note_fc_gp(engagement_fc=False, engagement_gp=False, part_gp=0, nb_personnes_gp=0):
    """
    Note financement collectif (max 2) ET gouvernance partagée (max 5).
    FC et GP sont non cumulables.
    """
    if engagement_gp:
        if part_gp > 0.5 and nb_personnes_gp >= 50:
            return 5, "GP max (>50% fonds propres, ≥50 personnes)"
        elif part_gp >= 0.40 and nb_personnes_gp >= 30:
            return 4, "GP niveau 2 (≥40%, ≥30 personnes)"
        elif part_gp >= 1/3 and nb_personnes_gp >= 20:
            return 3, "GP niveau 1 (≥1/3, ≥20 personnes)"
        else:
            return 0, "GP : conditions non remplies"
    elif engagement_fc:
        return 2, "FC engagé (10% financement local)"
    else:
        return 0, "Pas d'engagement FC/GP"


def simuler_note_totale(prix, ecs, periode=12, engagement_gp=False, part_gp=0,
                         nb_personnes_gp=0, engagement_fc=False):
    """Simule la note totale sur 100 points."""
    p_plafond = PRIX_PLAFOND.get(periode)
    np, msg_np = calculer_note_np(prix, p_plafond)
    nc, msg_nc = calculer_note_nc(ecs, periode)
    nfc, msg_nfc = calculer_note_fc_gp(engagement_fc, engagement_gp, part_gp, nb_personnes_gp)

    if np is None:
        np = 0  # simulation sans plafond connu

    note_totale = (np or 0) + nc + nfc
    return {
        "note_totale": round(note_totale, 2),
        "NP": {"note": np, "detail": msg_np, "poids": 70},
        "NC": {"note": nc, "detail": msg_nc, "poids": 25},
        "NFC": {"note": nfc, "detail": msg_nfc, "poids": 5},
        "periode": periode,
        "prix": prix,
        "ecs": ecs,
        "ecs_arrondi": round(ecs / 10) * 10 if ecs else None
    }


# ─── ROUTES ────────────────────────────────────────────────────────────────────

@ao_pv_bp.route('/')
def index():
    """Page principale - sélection du projet/prospect source."""
    try:
        from crm_routes import get_current_crm_user
        user_id, is_admin = get_current_crm_user()
    except Exception:
        # Fail-closed : en cas d'erreur on NE retombe JAMAIS sur admin.
        user_id, is_admin = None, False
    prospects = _get_prospects_for_ao(user_id=user_id, is_admin=is_admin)
    return render_template('ao_pv_batiment.html',
                           prospects=prospects,
                           periode=PERIODE_ACTUELLE,
                           date_limite=DATE_LIMITE_P12,
                           gwp_pays=list(GWP_PAR_PAYS.keys()),
                           carbone_ref=CARBONE_REF[PERIODE_ACTUELLE])


@ao_pv_bp.route('/api/prospect/<int:prospect_id>')
def get_prospect_data(prospect_id):
    """Retourne les données d'un prospect formatées pour le formulaire AO."""
    try:
        from database_adapter import execute_query
        from crm_routes import get_current_crm_user, verify_prospect_ownership
        # Isolation multi-tenant : exiger session + propriété du prospect.
        user_id, is_admin = get_current_crm_user()
        if user_id is None:
            return jsonify({"error": "Authentification requise"}), 401
        if not verify_prospect_ownership(prospect_id, user_id, is_admin):
            return jsonify({"error": "Prospect non trouvé"}), 404
        p = execute_query(
            'SELECT * FROM agriweb_prospects WHERE id = %s',
            (prospect_id,), fetch_one=True
        )
        if not p:
            return jsonify({"error": "Prospect non trouvé"}), 404

        # Surface toiture et puissance estimée
        surface_m2 = p.get('surface_m2')
        puissance_kwc_estimee = None

        # Essayer aussi dans data_json
        try:
            dj = p.get('data_json')
            if dj:
                if isinstance(dj, str):
                    dj = json.loads(dj)
                if isinstance(dj, dict):
                    surface_m2 = surface_m2 or dj.get('surface_toiture_m2') or dj.get('surface_m2')
        except Exception:
            pass

        if surface_m2:
            puissance_kwc_estimee = round(float(surface_m2) * 0.80 * 0.18, 0)

        data = {
            "id": p.get('id'),
            "raison_sociale": p.get('nom_prospect') or "",
            "adresse": p.get('adresse') or "",
            "commune": p.get('commune') or "",
            "code_postal": "",
            "departement": p.get('departement') or "",
            "contact_nom": p.get('contact_nom') or "",
            "contact_email": p.get('contact_email') or "",
            "contact_tel": p.get('contact_telephone') or "",
            "surface_toiture_m2": surface_m2,
            "puissance_kwc_estimee": puissance_kwc_estimee,
            "type": p.get('type') or "",
        }
        return jsonify(data)
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@ao_pv_bp.route('/api/calcul-carbone', methods=['POST'])
def calcul_carbone():
    """Calcule le bilan carbone simplifié."""
    data = request.get_json()
    pays_module = data.get("pays_module", "Chine")
    pays_cellule = data.get("pays_cellule", "Chine")
    pays_wafer = data.get("pays_wafer", "Chine")
    tech = data.get("technologie", "mono")

    G_arrondi, G_brut = estimation_carbone_rapide(pays_module, pays_cellule, pays_wafer, tech)

    periode = data.get("periode", PERIODE_ACTUELLE)
    ref = CARBONE_REF.get(periode, CARBONE_REF[PERIODE_ACTUELLE])
    nc, msg_nc = calculer_note_nc(G_arrondi, periode)

    return jsonify({
        "G_arrondi": G_arrondi,
        "G_brut": G_brut,
        "note_NC": nc,
        "detail_NC": msg_nc,
        "plafond": ref["sup"],
        "plancher": ref["inf"],
        "eligible": G_arrondi <= ref["sup"],
        "pays": {
            "module": pays_module,
            "cellule": pays_cellule,
            "wafer": pays_wafer
        }
    })


@ao_pv_bp.route('/api/simulation-notes', methods=['POST'])
def simulation_notes():
    """Simule les notes NP/NC/NFC et la note totale."""
    data = request.get_json()
    prix = float(data.get("prix", 0))
    ecs = float(data.get("ecs", 500))
    periode = int(data.get("periode", PERIODE_ACTUELLE))
    engagement_gp = data.get("engagement_gp", False)
    part_gp = float(data.get("part_gp", 0)) / 100  # % → fraction
    nb_personnes_gp = int(data.get("nb_personnes_gp", 0))
    engagement_fc = data.get("engagement_fc", False)

    result = simuler_note_totale(
        prix=prix,
        ecs=ecs,
        periode=periode,
        engagement_gp=engagement_gp,
        part_gp=part_gp,
        nb_personnes_gp=nb_personnes_gp,
        engagement_fc=engagement_fc
    )
    return jsonify(result)


@ao_pv_bp.route('/api/garantie-financiere', methods=['POST'])
def calcul_garantie():
    """Calcule le montant de la garantie financière de mise en œuvre (30 000 € × MWc)."""
    data = request.get_json()
    puissance_kwc = float(data.get("puissance_kwc", 0))
    puissance_mwc = puissance_kwc / 1000
    garantie = puissance_mwc * 30000
    demantelement = puissance_mwc * 10000 if data.get("agrivoltaique") else 0

    return jsonify({
        "puissance_mwc": round(puissance_mwc, 3),
        "garantie_mise_en_oeuvre_eur": garantie,
        "garantie_demantelement_eur": demantelement,
        "total_garanties_eur": garantie + demantelement
    })


@ao_pv_bp.route('/api/export-pdf', methods=['POST'])
def export_pdf():
    """Génère le formulaire de candidature pré-rempli en PDF."""
    data = request.get_json()
    try:
        pdf_buffer = _generer_pdf_formulaire(data)
        nom_projet = data.get("nom_projet", "projet").replace(" ", "_")
        return send_file(
            pdf_buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f"formulaire_AO_PV_P12_{nom_projet}_{datetime.now().strftime('%Y%m%d')}.pdf"
        )
    except Exception as e:
        return jsonify({"error": f"Erreur génération PDF : {str(e)}"}), 500


@ao_pv_bp.route('/api/export-excel', methods=['POST'])
def export_excel():
    """Génère le tableau de bilan carbone en Excel (.xlsx)."""
    data = request.get_json()
    try:
        excel_buffer = _generer_excel_carbone(data)
        nom_projet = data.get("nom_projet", "projet").replace(" ", "_")
        return send_file(
            excel_buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"bilan_carbone_AO_PV_P12_{nom_projet}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        )
    except Exception as e:
        return jsonify({"error": f"Erreur génération Excel : {str(e)}"}), 500


@ao_pv_bp.route('/api/checker-admissibilite', methods=['POST'])
def checker_admissibilite():
    """Vérifie les conditions d'admissibilité de l'offre."""
    data = request.get_json()
    problemes = []
    avertissements = []

    puissance = data.get("puissance_kwc", 0)
    ecs = data.get("ecs", 9999)
    ref = CARBONE_REF.get(PERIODE_ACTUELLE, {"sup": 740, "inf": 420})

    # Puissance > 500 kWc obligatoire
    if puissance <= 500:
        problemes.append(f"Puissance {puissance} kWc ≤ 500 kWc → Installation non éligible (seuil : > 500 kWc)")

    # Bilan carbone sous le plafond
    if ecs > ref["sup"]:
        problemes.append(f"ECS {ecs} kg CO2/kWc > plafond {ref['sup']} → Offre éliminée")

    # Autorisation d'urbanisme
    if not data.get("has_autorisation_urbanisme"):
        avertissements.append("Autorisation d'urbanisme manquante (PC ou DP + non-opposition)")

    # Garantie financière
    garantie_requise = (puissance / 1000) * 30000
    garantie_fournie = data.get("garantie_montant", 0)
    if garantie_fournie < garantie_requise:
        avertissements.append(f"Garantie financière insuffisante : {garantie_fournie:,.0f} € < {garantie_requise:,.0f} € requis")

    # Certifications ISO
    if not data.get("iso_9001_modules"):
        avertissements.append("Certification ISO 9001 fabricant modules à vérifier")
    if not data.get("iso_14001_modules"):
        avertissements.append("Certification ISO 14001 fabricant modules à vérifier")

    eligible = len(problemes) == 0

    return jsonify({
        "eligible": eligible,
        "problemes": problemes,
        "avertissements": avertissements,
        "score_conformite": max(0, 100 - len(problemes) * 30 - len(avertissements) * 5)
    })


# ─── HELPERS ───────────────────────────────────────────────────────────────────

def _get_prospects_for_ao(user_id=None, is_admin=False):
    """Récupère les prospects depuis agriweb_prospects via database_adapter."""
    try:
        from database_adapter import execute_query
        from crm_routes import user_filter_clause
        # Fail-closed : un utilisateur non résolu ne voit AUCUN prospect
        # (auparavant on vidait le filtre -> liste de TOUS les tenants).
        if not is_admin and user_id is None:
            return []
        filter_clause, params = user_filter_clause(user_id, is_admin)

        rows = execute_query(
            f'SELECT id, nom_prospect, commune, departement, adresse, statut, surface_m2 '
            f'FROM agriweb_prospects WHERE 1=1{filter_clause} '
            f'ORDER BY date_creation DESC LIMIT 200',
            params if params else None, fetch_all=True
        )

        result = []
        for p in (rows or []):
            result.append({
                "id": p.get('id'),
                "company_name": p.get('nom_prospect') or '(sans nom)',
                "city": p.get('commune') or '',
                "department": p.get('departement') or '',
                "status": p.get('statut') or 'nouveau',
                "address": p.get('adresse') or '',
            })
        return result
    except Exception as e:
        print(f"[AO PV] Impossible de charger les prospects: {e}")
        import traceback; traceback.print_exc()
        return []


def _generer_pdf_formulaire(data):
    """Génère le PDF du formulaire de candidature AO pré-rempli."""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import cm
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
    except ImportError:
        # Fallback texte si reportlab non dispo
        return _generer_pdf_fallback(data)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()

    ORANGE = colors.HexColor('#F7971E')
    DARK = colors.HexColor('#1a1a2e')
    LIGHT_GREY = colors.HexColor('#f5f5f5')

    style_titre = ParagraphStyle('titre', fontSize=14, textColor=DARK, spaceAfter=6,
                                  alignment=TA_CENTER, fontName='Helvetica-Bold')
    style_sous_titre = ParagraphStyle('sous_titre', fontSize=11, textColor=ORANGE, spaceAfter=4,
                                       fontName='Helvetica-Bold')
    style_normal = ParagraphStyle('normal', fontSize=9, spaceAfter=3)
    style_note = ParagraphStyle('note', fontSize=8, textColor=colors.grey, spaceAfter=2, leftIndent=10)

    story = []

    # En-tête
    story.append(Paragraph("FORMULAIRE DE CANDIDATURE", style_titre))
    story.append(Paragraph("Appel d'offres PPE2 PV Bâtiment – Centrales sur bâtiments > 500 kWc", style_titre))
    story.append(Paragraph(f"Période 12 – Date limite : {DATE_LIMITE_P12}", style_normal))
    story.append(HRFlowable(width="100%", thickness=2, color=ORANGE))
    story.append(Spacer(1, 0.3*cm))

    def section(titre, champs):
        story.append(Paragraph(titre, style_sous_titre))
        table_data = []
        for label, valeur in champs:
            table_data.append([
                Paragraph(f"<b>{label}</b>", style_normal),
                Paragraph(str(valeur) if valeur else "—", style_normal)
            ])
        t = Table(table_data, colWidths=[6*cm, 11*cm])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), LIGHT_GREY),
            ('TEXTCOLOR', (0, 0), (-1, -1), DARK),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.white, LIGHT_GREY]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        story.append(t)
        story.append(Spacer(1, 0.3*cm))

    # Section A - Candidat
    section("A. IDENTIFICATION DU CANDIDAT", [
        ("Raison sociale", data.get("raison_sociale")),
        ("Représentant légal", data.get("representant_legal")),
        ("Adresse", data.get("adresse")),
        ("Code postal", data.get("code_postal")),
        ("Commune", data.get("commune")),
        ("Téléphone", data.get("contact_tel")),
        ("Email", data.get("contact_email")),
        ("SIRET", data.get("siret")),
    ])

    # Section B - Installation
    section("B. DESCRIPTION DE L'INSTALLATION", [
        ("Nom du projet", data.get("nom_projet")),
        ("Type d'installation", data.get("type_installation", "Bâtiment")),
        ("Adresse du site", data.get("adresse_site")),
        ("Commune du site", data.get("commune_site")),
        ("Département", data.get("departement")),
        ("Puissance installée (kWc)", data.get("puissance_kwc")),
        ("Surface toiture (m²)", data.get("surface_toiture_m2")),
        ("Nature d'exploitation", data.get("nature_exploitation", "Vente totale")),
        ("Autoconsommation (%)", data.get("taux_autoconso", "0")),
        ("Technologie modules", data.get("technologie_module", "Silicium monocristallin")),
        ("Fabricant modules", data.get("fabricant_module")),
        ("Pays fabrication module", data.get("pays_module")),
        ("Pays fabrication cellule", data.get("pays_cellule")),
        ("Pays fabrication wafer", data.get("pays_wafer")),
        ("Référence autorisation urbanisme", data.get("ref_autorisation")),
    ])

    # Section C - Offre financière
    section("C. OFFRE FINANCIÈRE", [
        ("Prix de référence T₀ (€/MWh)", data.get("prix_reference")),
        ("Indexation K souhaitée", "Oui" if data.get("indexation_k") else "Non (par défaut)"),
        ("Évaluation carbone simplifiée (kg CO₂/kWc)", data.get("ecs_valeur")),
        ("Période de candidature", f"Période {PERIODE_ACTUELLE} – {DATE_LIMITE_P12}"),
        ("Financement collectif", "Oui" if data.get("engagement_fc") else "Non"),
        ("Gouvernance partagée", "Oui" if data.get("engagement_gp") else "Non"),
    ])

    # Notes simulées
    if data.get("simulation"):
        sim = data["simulation"]
        story.append(Paragraph("SIMULATION DE LA NOTE (indicative)", style_sous_titre))
        note_data = [
            ["Critère", "Note", "Pondération", "Commentaire"],
            ["NP (Prix)", str(sim.get("NP", {}).get("note", "—")), "70 pts", sim.get("NP", {}).get("detail", "")],
            ["NC (Carbone)", str(sim.get("NC", {}).get("note", "—")), "25 pts", sim.get("NC", {}).get("detail", "")],
            ["NFC (Gov./Fin.)", str(sim.get("NFC", {}).get("note", "—")), "5 pts", sim.get("NFC", {}).get("detail", "")],
            ["TOTAL", str(sim.get("note_totale", "—")), "100 pts", "Simulation indicative"],
        ]
        t = Table(note_data, colWidths=[3.5*cm, 2*cm, 2.5*cm, 9*cm])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), ORANGE),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ('BACKGROUND', (0, -1), (-1, -1), DARK),
            ('TEXTCOLOR', (0, -1), (-1, -1), colors.white),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('LEFTPADDING', (0, 0), (-1, -1), 5),
            ('RIGHTPADDING', (0, 0), (-1, -1), 5),
        ]))
        story.append(t)
        story.append(Spacer(1, 0.3*cm))

    # Garanties financières
    puissance_kwc = float(data.get("puissance_kwc", 0) or 0)
    if puissance_kwc > 0:
        garantie = (puissance_kwc / 1000) * 30000
        section("D. GARANTIES FINANCIÈRES", [
            ("Montant garantie mise en œuvre", f"{garantie:,.0f} €"),
            ("Calcul", f"30 000 € × {puissance_kwc/1000:.3f} MWc"),
            ("Forme", data.get("forme_garantie", "Garantie bancaire à 1ère demande")),
        ])

    # Pied de page
    story.append(HRFlowable(width="100%", thickness=1, color=colors.lightgrey))
    story.append(Paragraph(
        f"Document généré par HeliaPV – {datetime.now().strftime('%d/%m/%Y %H:%M')} – "
        f"AO PPE2 PV Bâtiment, Appel d'offres CRE, Période {PERIODE_ACTUELLE}",
        style_note
    ))
    story.append(Paragraph(
        "⚠ Ce document est une aide à la préparation. Vérifiez la conformité avec le CDC officiel disponible sur cre.fr",
        ParagraphStyle('warning', fontSize=8, textColor=colors.red, spaceAfter=2)
    ))

    doc.build(story)
    buffer.seek(0)
    return buffer


def _generer_pdf_fallback(data):
    """PDF basique si reportlab non disponible."""
    content = f"""FORMULAIRE AO PV BÂTIMENT - PÉRIODE {PERIODE_ACTUELLE}
Généré le {datetime.now().strftime('%d/%m/%Y %H:%M')}

CANDIDAT: {data.get('raison_sociale', 'N/A')}
COMMUNE: {data.get('commune', 'N/A')}
PUISSANCE: {data.get('puissance_kwc', 'N/A')} kWc
PRIX T0: {data.get('prix_reference', 'N/A')} €/MWh
ECS: {data.get('ecs_valeur', 'N/A')} kg CO2/kWc
"""
    buffer = io.BytesIO(content.encode('utf-8'))
    buffer.seek(0)
    return buffer


def _generer_excel_carbone(data):
    """Génère le tableau 1 du bilan carbone simplifié en Excel."""
    try:
        import openpyxl
        from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
    except ImportError:
        raise Exception("openpyxl non installé. Installez-le avec: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bilan Carbone Simplifié"

    # Styles
    orange_fill = PatternFill("solid", fgColor="F7971E")
    dark_fill = PatternFill("solid", fgColor="1a1a2e")
    grey_fill = PatternFill("solid", fgColor="f0f0f0")
    bold_font = Font(bold=True)
    white_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Titre
    ws.merge_cells('A1:G1')
    ws['A1'] = f"TABLEAU 1 - Bilan Carbone Simplifié AO PPE2 PV Bâtiment – Période {PERIODE_ACTUELLE}"
    ws['A1'].font = Font(bold=True, size=12, color="FFFFFF")
    ws['A1'].fill = dark_fill
    ws['A1'].alignment = center_align

    ws.merge_cells('A2:G2')
    ws['A2'] = f"Projet : {data.get('nom_projet', '—')} | Candidat : {data.get('raison_sociale', '—')} | Date : {datetime.now().strftime('%d/%m/%Y')}"
    ws['A2'].font = Font(bold=True, size=9)
    ws['A2'].fill = grey_fill
    ws['A2'].alignment = center_align

    # En-têtes tableau Annexe 2
    headers = [
        "Composant", "Quantité/kWc\n(Qi)", "Unité",
        "Pays fabrication\n(site j)", "Fraction\n(xij)",
        "GWP unitaire\n(kg CO2/unité)", "Gi\n(kg CO2/kWc)"
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = white_font
        cell.fill = orange_fill
        cell.alignment = center_align
        cell.border = border

    # Données composants
    tech = data.get("technologie_module", "mono").lower()
    if "mono" in tech:
        tech_key = "mono"
    elif "multi" in tech:
        tech_key = "multi"
    else:
        tech_key = "mono"

    pays_m = data.get("pays_module", "Autre pays du Monde")
    pays_c = data.get("pays_cellule", pays_m)
    pays_w = data.get("pays_wafer", pays_c)

    gwp_m = GWP_PAR_PAYS.get(pays_m, GWP_PAR_PAYS["Autre pays du Monde"])
    gwp_c = GWP_PAR_PAYS.get(pays_c, GWP_PAR_PAYS["Autre pays du Monde"])
    gwp_w = GWP_PAR_PAYS.get(pays_w, GWP_PAR_PAYS["Autre pays du Monde"])

    # Quantités standard pour module 400Wc (simplification pour tableau de base)
    surface_kwc = 2.5  # m²/kWc
    composants_data = [
        ("Module cristallin", round(surface_kwc, 3), "m²/kWc", pays_m, 1.0, gwp_m.get("module_cristallin", 6.5), round(surface_kwc * gwp_m.get("module_cristallin", 6.5), 2)),
        ("Cellule", round(surface_kwc * 1.02, 3), "m²/kWc", pays_c, 1.0, gwp_c.get("cellule", 22.0), round(surface_kwc * 1.02 * gwp_c.get("cellule", 22.0), 2)),
        (f"Plaquette wafer ({tech_key})", round(surface_kwc * 1.02 * 1.01, 3), "m²/kWc", pays_w, 1.0, gwp_w.get(f"wafer_{tech_key}", 5.0), round(surface_kwc * 1.02 * 1.01 * gwp_w.get(f"wafer_{tech_key}", 5.0), 2)),
        ("Verre (avant)", round(surface_kwc * 0.8, 3), "kg/kWc", pays_m, 1.0, gwp_m.get("verre", 1.0), round(surface_kwc * 0.8 * gwp_m.get("verre", 1.0), 2)),
        ("Encapsulant EVA", round(surface_kwc * 0.5 * 1.01, 3), "kg/kWc", pays_m, 1.0, gwp_m.get("encapsulant", 2.7), round(surface_kwc * 0.5 * 1.01 * gwp_m.get("encapsulant", 2.7), 2)),
        ("Face arrière PET", round(surface_kwc * 0.3 * 1.02, 3), "kg/kWc", pays_m, 1.0, gwp_m.get("face_arriere", 3.7), round(surface_kwc * 0.3 * 1.02 * gwp_m.get("face_arriere", 3.7), 2)),
    ]

    G_total = 0
    for row_idx, row_data in enumerate(composants_data, 5):
        fill = grey_fill if row_idx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = border
            cell.fill = fill
            cell.alignment = Alignment(horizontal='center' if col > 1 else 'left', vertical='center')
        G_total += row_data[6]

    # Ligne total
    last_row = 5 + len(composants_data)
    ws.cell(row=last_row, column=1, value="TOTAL G (kg CO2-eq/kWc)").font = white_font
    ws.cell(row=last_row, column=1).fill = dark_fill
    ws.cell(row=last_row, column=1).border = border
    G_arrondi = round(G_total / 10) * 10
    ws.cell(row=last_row, column=7, value=f"{G_arrondi} (arrondi)").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=last_row, column=7).fill = dark_fill
    ws.cell(row=last_row, column=7).border = border
    for col in range(2, 7):
        cell = ws.cell(row=last_row, column=col, value=round(G_total, 1) if col == 6 else "")
        cell.fill = dark_fill
        cell.border = border

    # Comparaison avec les seuils
    ws.cell(row=last_row + 2, column=1, value="Validation").font = bold_font
    ref = CARBONE_REF[PERIODE_ACTUELLE]
    ws.cell(row=last_row + 2, column=2, value=f"Plafond ECS : {ref['sup']} kg CO2/kWc")
    ws.cell(row=last_row + 3, column=2, value=f"Plancher ECS : {ref['inf']} kg CO2/kWc")
    status = "✓ ELIGIBLE" if G_arrondi <= ref["sup"] else "✗ HORS PLAFOND"
    ws.cell(row=last_row + 4, column=2, value=f"Statut : {status} ({G_arrondi} kg CO2/kWc)")
    ws.cell(row=last_row + 4, column=2).font = Font(bold=True, color="00AA00" if "ELIGIBLE" in status else "FF0000")

    # Largeurs colonnes
    widths = [25, 14, 10, 22, 10, 18, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # Onglet Pays de référence
    ws2 = wb.create_sheet("GWP Pays de référence")
    headers2 = ["Pays", "Module cristallin\nm²", "Cellule\nkgCO2/m²", "Wafer mono\nm²", "Verre\nkg", "Encapsulant\nkg", "Face arrière\nkg"]
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = white_font
        cell.fill = dark_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    for row_i, (pays, gwp) in enumerate(GWP_PAR_PAYS.items(), 2):
        ws2.cell(row=row_i, column=1, value=pays)
        ws2.cell(row=row_i, column=2, value=gwp.get("module_cristallin"))
        ws2.cell(row=row_i, column=3, value=gwp.get("cellule"))
        ws2.cell(row=row_i, column=4, value=gwp.get("wafer_mono"))
        ws2.cell(row=row_i, column=5, value=gwp.get("verre"))
        ws2.cell(row=row_i, column=6, value=gwp.get("encapsulant"))
        ws2.cell(row=row_i, column=7, value=gwp.get("face_arriere"))
        for col in range(1, 8):
            ws2.cell(row=row_i, column=col).border = border

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
