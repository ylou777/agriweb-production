"""Script de génération de la base de données Certisolis (JSON).
À relancer à chaque mise à jour des fichiers Certisolis.
"""
import openpyxl, json
from datetime import datetime

KEY_VALIDE = "ECS Valable aujourd'hui"

def norm_key(k):
    return str(k).strip().replace('\n', ' ')

def parse_sheet(wb, sn, method):
    ws = wb[sn]
    rows = list(ws.iter_rows(values_only=True))
    header = [norm_key(h) if h else None for h in rows[3]]
    data = []
    for row in rows[4:]:
        if row[1] is None:
            break
        r = {'methode': method}
        for j, h in enumerate(header):
            if h and j < len(row):
                val = row[j]
                if hasattr(val, 'strftime'):
                    val = val.strftime('%Y-%m-%d')
                if isinstance(val, str):
                    val = val.strip()
                r[h] = val
        r['fabricant_norm'] = str(r.get('Nom du fabricant', '') or '').upper().strip()
        r['refs_norm'] = str(r.get('Références modules', '') or '').upper().strip()
        data.append(r)
    return data

wb1 = openpyxl.load_workbook('certisolis_ppe2.xlsx', read_only=True, data_only=True)
wb2 = openpyxl.load_workbook('certisolis_ppe2v2.xlsx', read_only=True, data_only=True)

ppe2   = parse_sheet(wb1, 'ECS PPE2 en vigueur', 'PPE2')
ppe2v2 = parse_sheet(wb2, 'ECS PPE2_V2', 'PPE2-V2')

all_data = ppe2 + ppe2v2
valides = [r for r in all_data if r.get(KEY_VALIDE) == 'OUI']
valides_v2 = [r for r in ppe2v2 if r.get(KEY_VALIDE) == 'OUI']

print(f'Total lignes : {len(all_data)}')
print(f'Valides (toutes méthodes) : {len(valides)}')
print(f'Valides PPE2-V2 : {len(valides_v2)}')
print(f'Fabricants PPE2-V2 valides : {sorted(set(r["fabricant_norm"] for r in valides_v2))}')

with open('certisolis_db.json', 'w', encoding='utf-8') as f:
    json.dump({
        'generated': datetime.now().strftime('%Y-%m-%d'),
        'source_ppe2':    'Tableau-du-10-02-2026-PPE2.xlsx',
        'source_ppe2v2':  'Tableau-du-10-02-2026-PPE2_V2.xlsx',
        'total': len(all_data),
        'data': all_data
    }, f, ensure_ascii=False, indent=2)
print('certisolis_db.json written.')
