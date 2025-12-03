import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Lire le fichier CSV
input_file = 'societes_pv_enrichies.csv'
output_file = 'societes_pv_a_completer.xlsx'

with open(input_file, 'r', encoding='utf-8') as f:
    reader = csv.DictReader(f)
    donnees = list(reader)

# Créer un workbook Excel
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Societes PV"

# Headers
headers = ['Société Recherchée', 'SIREN', 'Dénomination', 'Téléphone', 'Email', 'Contact', 'Remarques']
ws.append(headers)

# Styliser les headers
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=12)

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Ajouter les données
for row in donnees:
    societe = row['societe_recherchee']
    siren = row['siren']
    denomination = row['denomination']
    
    # Déterminer le statut
    if siren == 'NON TROUVE':
        telephone = 'À rechercher'
        email = 'À rechercher'
        contact = 'À rechercher'
    else:
        telephone = ''  # À compléter manuellement
        email = ''      # À compléter manuellement
        contact = ''    # À compléter manuellement
    
    ws.append([societe, siren, denomination, telephone, email, contact, ''])

# Ajuster la largeur des colonnes
ws.column_dimensions['A'].width = 35
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 45
ws.column_dimensions['D'].width = 18
ws.column_dimensions['E'].width = 30
ws.column_dimensions['F'].width = 25
ws.column_dimensions['G'].width = 40

# Colorer les lignes "NON TROUVE" en rouge clair
red_fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
green_fill = PatternFill(start_color='E6FFE6', end_color='E6FFE6', fill_type='solid')

for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
    siren_cell = row[1]
    if siren_cell.value == 'NON TROUVE':
        for cell in row:
            cell.fill = red_fill
    else:
        for cell in row:
            cell.fill = green_fill

# Figer la première ligne
ws.freeze_panes = 'A2'

# Sauvegarder
wb.save(output_file)

print(f"Fichier Excel créé: {output_file}")
print(f"Total sociétés: {len(donnees)}")
print(f"SIREN trouvés: {len([d for d in donnees if d['siren'] != 'NON TROUVE'])}")
print(f"Non trouvés: {len([d for d in donnees if d['siren'] == 'NON TROUVE'])}")
print(f"\nLes colonnes Téléphone, Email et Contact sont prêtes à être complétées!")
print(f"Astuce: Utilisez https://www.pappers.fr ou https://www.societe.com avec les SIREN")
